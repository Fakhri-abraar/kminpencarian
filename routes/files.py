# files.py

"""
File management routes
Handles file upload, download, listing, and deletion
"""
from flask import (
    Blueprint, render_template, request, redirect, url_for, flash, 
    send_file, make_response, send_from_directory, current_app
)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
import os
import io
import time
import uuid
import subprocess
import sys
from datetime import datetime

import openpyxl
from io import BytesIO
import base64

from extensions import db
from models.file import File
from models.report import FinancialReport
from models.user import User
from werkzeug.exceptions import BadRequest
from models.access import UserAccess
from models.connection import Connection
from models.file_access_request import FileAccessRequest
from cryptography.hazmat.primitives import serialization
from utils.rsa_handler import load_public_key, encrypt_with_public_key
from utils.nosql_handler import get_user_public_key, store_file_key
from utils.rsa_handler import load_private_key, decrypt_with_private_key
from utils.nosql_handler import get_file_key, get_shared_key, get_user_private_key_enc
from encryption.aes_handler import AESHandler
from encryption.des_handler import DESHandler
from encryption.rc4_handler import RC4Handler

from utils.key_manager import generate_file_key, encrypt_file_key, decrypt_file_key
from utils.pbe_handler import derive_key_from_password


from utils.file_handler import (
    is_allowed_file, generate_unique_filename, get_file_size,
    validate_file_size, save_encrypted_file, read_encrypted_file,
    format_file_size, get_file_category, ensure_upload_directory
)
from utils.validators import (
    validate_algorithm, validate_filename, 
    validate_encryption_password_length # <-- IMPOR BARU
)
from utils.logger import log_crypto_operation

files_bp = Blueprint('files', __name__, url_prefix='/files')

ensure_upload_directory()

# --- FUNGSI PARSING EXCEL (Tidak Berubah) ---
def create_parsed_excel(file_data, handler):
    try:
        workbook = openpyxl.load_workbook(BytesIO(file_data))
        sheet = workbook.active
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    try:
                        data_to_encrypt = str(cell.value).encode('utf-8')
                        ciphertext, iv, _ = handler.encrypt(data_to_encrypt)
                        combined_data = (iv if iv else b'') + ciphertext
                        encrypted_string = base64.b64encode(combined_data).decode('utf-8')
                        cell.value = encrypted_string
                    except Exception as e:
                        cell.value = f"Error: {str(e)}"
        output_bio = BytesIO()
        workbook.save(output_bio)
        return output_bio.getvalue()
    except Exception as e:
        print(f"Gagal memproses excel per-sel: {e}")
        return None

# --- RUTE UPLOAD (Perubahan di sini) ---
@files_bp.route('/upload', methods=['GET', 'POST'])
@login_required
def upload():
    # 1. Cek Role: Hanya Organization yang boleh upload
    if current_user.role != 'organization':
        flash('Only organizations can upload files.', 'error')
        return redirect(url_for('main.dashboard'))

    if request.method == 'GET':
        return render_template('upload.html')
    
    file = request.files['file']
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('files.upload'))
    
    # Ambil algoritma (Default AES)
    algorithm = request.form.get('algorithm', 'AES').upper()
    
    # Validasi File (Tetap dipertahankan)
    is_valid_name, sanitized_name, name_error = validate_filename(file.filename)
    if not is_valid_name: 
        flash(name_error, 'error')
        return redirect(url_for('files.upload'))
        
    if not is_allowed_file(sanitized_name): 
        flash('File type not allowed', 'error')
        return redirect(url_for('files.upload'))
        
    is_valid_size, size_error = validate_file_size(file)
    if not is_valid_size: 
        flash(size_error, 'error')
        return redirect(url_for('files.upload'))
    
    try:
        # --- [MODIFIKASI TAHAP 2: Hybrid Encryption] ---
        
        file_data = file.read()
        file_size = len(file_data)
        unique_filename = generate_unique_filename(sanitized_name)

        # 1. Generate Random Symmetric Key (32 bytes untuk keamanan maksimal)
        # Kunci ini digenerate sistem, bukan dari password user lagi
        file_key = os.urandom(32)
        
        # 2. Setup Handler Enkripsi (AES/DES/RC4) menggunakan Random Key
        if algorithm == 'AES':
            handler = AESHandler(file_key)
        elif algorithm == 'DES':
            # DES hanya menggunakan 8 byte pertama
            handler = DESHandler(file_key[:8])
        elif algorithm == 'RC4':
            # RC4 kita gunakan 16 byte (128-bit)
            handler = RC4Handler(file_key[:16])
        else:
            flash('Invalid algorithm', 'error')
            return redirect(url_for('files.upload'))
        
        # 3. Enkripsi File Fisik
        start_time = time.time()
        ciphertext, iv, encryption_time = handler.encrypt(file_data)
        
        # 4. Ambil Public Key Owner (Organization) dari MongoDB
        # Ini memastikan hanya owner yang bisa membuka kunci ini nanti (via Private Key-nya)
        user_pub_key_pem = get_user_public_key(current_user.id)
        if not user_pub_key_pem:
            raise Exception("Public Key not found. Please contact admin to generate keys.")
            
        public_key = load_public_key(user_pub_key_pem)
        
        # 5. Enkripsi Symmetric Key (file_key) dengan RSA Public Key
        # Kita mengenkripsi full 32 bytes key master
        encrypted_file_key = encrypt_with_public_key(public_key, file_key)
        
        # 6. Simpan File Fisik Terenkripsi ke Disk
        save_encrypted_file(ciphertext, unique_filename)
        
        # 7. Proses Excel (Parsing) jika tipe file Excel
        # Handler sudah menggunakan kunci baru, jadi fungsi ini tetap aman
        parsed_filename = None
        if get_file_category(sanitized_name) == 'excel':
            parsed_excel_data = create_parsed_excel(file_data, handler) 
            if parsed_excel_data:
                parsed_filename = f"parsed_{unique_filename}"
                save_encrypted_file(parsed_excel_data, parsed_filename)
        
        # 8. Simpan Metadata ke MySQL
        # Catatan: Kolom 'salt' diisi dummy random karena tidak lagi dipakai untuk derivasi kunci,
        # tapi kolom database tidak boleh kosong (non-nullable).
        file_record = File(
            file_uuid=str(uuid.uuid4()),
            owner_id=current_user.id,
            original_filename=sanitized_name,
            encrypted_filename=unique_filename,
            parsed_filename=parsed_filename,
            file_size=file_size,
            encrypted_size=len(ciphertext),
            file_type=get_file_category(sanitized_name),
            encryption_algorithm=algorithm,
            cipher_mode='CBC' if algorithm in ['AES', 'DES'] else None,
            salt=os.urandom(16).hex(), # Dummy Salt
            iv=iv.hex() if iv else None,
            encryption_time=encryption_time,
            uploaded_at=datetime.utcnow(),
            upload_date=datetime.utcnow()
        )
        
        db.session.add(file_record)
        db.session.commit() # Commit untuk mendapatkan ID file
        
        # 9. Simpan Encrypted Symmetric Key ke MongoDB
        # Kita simpan kuncinya di NoSQL agar terpisah dari database metadata utama
        store_file_key(file_record.id, current_user.id, encrypted_file_key)
        
        # 10. Logging
        log_crypto_operation(
            user_id=current_user.id, file_id=file_record.id, operation_type='encryption',
            algorithm=algorithm, file_size=file_size, execution_time=encryption_time, success=True
        )
        
        # 11. Proses Data Finansial (Excel) ke Database
        if get_file_category(sanitized_name) == 'excel':
            try:
                process_excel_file(file_data, file_record.id, handler)
                flash(f'File uploaded and encrypted successfully with {algorithm} (Hybrid)! Excel data processed.', 'success')
            except Exception as excel_error:
                flash(f'File uploaded with {algorithm}, but Excel processing failed: {str(excel_error)}', 'info')
        else:
            flash(f'File uploaded and encrypted successfully with {algorithm} (Hybrid)!', 'success')
        
        return redirect(url_for('main.dashboard'))
        
    except BadRequest as e:
        flash(str(e), 'error')
        return redirect(url_for('files.upload'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error uploading file: {str(e)}', 'error')
        return redirect(url_for('files.upload'))


@files_bp.route('/my-files')
@login_required
def my_files():
    """Menampilkan file milik pengguna yang sedang login."""
    files = File.query.filter_by(owner_id=current_user.id).order_by(File.upload_date.desc()).all()
    for file in files:
        file.formatted_size = format_file_size(file.file_size)
    return render_template('files.html', files=files)

# --- RUTE BARU UNTUK MELIHAT FILE PENGGUNA LAIN ---
@files_bp.route('/user/<string:username>')
@login_required
def view_user_files(username):
    """Menampilkan halaman profil file pengguna lain (User A)"""
    profile_user = User.query.filter_by(username=username).first_or_404()
    
    if profile_user.id == current_user.id:
        return redirect(url_for('files.my_files'))

    can_view = False
    
    if not profile_user.is_private:
        can_view = True
    else:
        # Check if there's a UserAccess record
        access_record = UserAccess.query.filter_by(
            owner_id=profile_user.id,
            authorized_user_id=current_user.id
        ).first()
        if connection:
            can_view = True
        
        # Also check if there's an accepted connection
        if not can_view:
            connection = Connection.query.filter(
                ((Connection.requester_id == profile_user.id) & (Connection.receiver_id == current_user.id)) |
                ((Connection.requester_id == current_user.id) & (Connection.receiver_id == profile_user.id)),
                Connection.status == 'accepted'
            ).first()
            if connection:
                can_view = True

    if can_view:
        from models.file_access_request import FileAccessRequest
        
        files = profile_user.files.order_by(File.upload_date.desc()).all()
        for file in files:
            file.formatted_size = format_file_size(file.file_size)
        
        # Check which files the user has been granted access to
        granted_file_ids = []
        approved_requests = FileAccessRequest.query.filter_by(
            requester_id=current_user.id,
            owner_id=profile_user.id,
            status='approved'
        ).all()
        
        for req in approved_requests:
            if req.file_id:
                granted_file_ids.append(req.file_id)
            else:
                # If file_id is None, they have access to all files
                granted_file_ids = [f.id for f in files]
                break
        
        # Check which files have pending access requests
        pending_file_ids = []
        pending_requests = FileAccessRequest.query.filter_by(
            requester_id=current_user.id,
            owner_id=profile_user.id,
            status='pending'
        ).all()
        
        for req in pending_requests:
            if req.file_id:
                pending_file_ids.append(req.file_id)
            else:
                # If file_id is None, they requested access to all files
                pending_file_ids = [f.id for f in files]
                break
        
        return render_template('user_files.html', files=files, profile_user=profile_user, is_private=False, granted_file_ids=granted_file_ids, pending_file_ids=pending_file_ids)
    else:
        # Instead of redirecting, render the user profile page and show a private account message
        return render_template('user_files.html', files=[], profile_user=profile_user, is_private=True, granted_file_ids=[], pending_file_ids=[])
# --- AKHIR RUTE BARU ---


# --- MODIFIKASI HELPER AKSES ---
def user_can_access_file(file_id, user_id):
    """
    Memeriksa apakah pengguna (user_id) dapat mengakses file (file_id) berdasarkan koneksi atau file access request.
    """
    file_record = db.session.get(File, file_id)
    if not file_record:
        return False, None
    
    if file_record.owner_id == user_id:
        return True, file_record

    # Check if there's an approved file access request for this specific file
    file_access_request = FileAccessRequest.query.filter_by(
        requester_id=user_id,
        owner_id=file_record.owner_id,
        file_id=file_id,
        status='approved'
    ).first()
    
    if file_access_request:
        return True, file_record

    # Check if there's an accepted connection between the users
    connection = Connection.query.filter(
        ((Connection.requester_id == file_record.owner_id) & (Connection.receiver_id == user_id)) |
        ((Connection.requester_id == user_id) & (Connection.receiver_id == file_record.owner_id)),
        Connection.status == 'accepted'
    ).first()
    
    if connection:
        return True, file_record
        
    return False, None
# --- AKHIR MODIFIKASI HELPER ---


# --- RUTE DOWNLOAD HIBRIDA (Tidak Berubah dari Langkah 1) ---
@files_bp.route('/decrypt/<int:file_id>', methods=['GET'])
@login_required
def download_logic(file_id):
    """
    Display decryption page for any file access, requiring password entry
    """
    can_access, file_record = user_can_access_file(file_id, current_user.id)
    
    if not can_access:
        flash('You do not have permission to access this file', 'error')
        return redirect(url_for('main.dashboard'))
    
    return render_template('decrypt.html', file=file_record)


@files_bp.route('/decrypt/<int:file_id>', methods=['POST'])
@login_required
def handle_download(file_id):
    """
    Handle decryption for both Owner and Consultant using RSA Keys.
    Menggantikan logika lama yang berbasis Password-Based Encryption (PBE).
    """
    # 1. Cek izin akses dasar (Owner atau Shared)
    can_access, file_record = user_can_access_file(file_id, current_user.id)
    
    # Cek referrer untuk redirect yang tepat jika error
    referrer = request.referrer
    from_user_files = referrer and '/files/user/' in referrer
    from_dashboard = referrer and '/dashboard' in referrer
    
    if not can_access:
        flash('You do not have permission to access this file', 'error')
        if from_user_files or from_dashboard:
            return redirect(referrer)
        return redirect(url_for('main.dashboard'))

    # 2. Ambil Password Login dari Form
    password = request.form.get('password')
    if not password:
        flash('Login password is required to decrypt your private key', 'error')
        if from_user_files or from_dashboard:
            return redirect(referrer)
        return redirect(url_for('files.download_logic', file_id=file_id))

    try:
        # 3. Ambil Encrypted Private Key User Saat Ini dari MongoDB
        # Kunci ini diperlukan untuk membuka identitas digital user
        user_priv_enc = get_user_private_key_enc(current_user.id)
        if not user_priv_enc:
            raise Exception("Your private key verification failed. Keys not found.")

        # 4. Decrypt Private Key User pakai Password Login
        # Jika password salah, proses ini akan gagal (ValueError)
        user_private_key = load_private_key(user_priv_enc, password)

        # 5. Tentukan sumber kunci file (Apakah saya Owner atau Konsultan?)
        encrypted_file_key = None
        
        if file_record.owner_id == current_user.id:
            # Jika saya Owner, ambil kunci dari koleksi 'file_keys'
            encrypted_file_key = get_file_key(file_id, current_user.id)
        else:
            # Jika saya Konsultan (Diberi Akses), ambil dari 'shared_keys'
            encrypted_file_key = get_shared_key(file_id, current_user.id)
            
        if not encrypted_file_key:
            raise Exception("Decryption key not found for your account. Please request access again.")

        # 6. Decrypt File Key (RSA) -> Menjadi RAW AES/DES/RC4 KEY
        # Membuka 'bungkusan' kunci menggunakan Private Key user
        raw_file_key = decrypt_with_private_key(user_private_key, encrypted_file_key)

        # 7. Lanjut ke proses dekripsi file fisik menggunakan Raw Key
        return decrypt_file_data_v2(file_record, raw_file_key, current_user.id)
        
    except ValueError:
        # Error ini biasanya muncul dari cryptography jika password salah
        flash('Invalid login password. Cannot unlock your private key.', 'error')
        if from_user_files or from_dashboard:
            return redirect(referrer)
        return redirect(url_for('files.download_logic', file_id=file_id))
        
    except Exception as e:
        flash(f'Decryption failed: {str(e)}', 'error')
        if from_user_files or from_dashboard:
            return redirect(referrer)
        return redirect(url_for('files.download_logic', file_id=file_id))


def decrypt_file_data(file_record, file_key, user_id):
    """Fungsi helper terpusat untuk dekripsi (Tidak Berubah)"""
    try:
        encrypted_data = read_encrypted_file(file_record.encrypted_filename)
        iv = bytes.fromhex(file_record.iv) if file_record.iv else None
        algorithm = file_record.encryption_algorithm

        if algorithm == 'AES': handler = AESHandler(file_key)
        elif algorithm == 'DES': handler = DESHandler(file_key)
        elif algorithm == 'RC4': handler = RC4Handler(file_key)
        else: raise Exception("Unknown algorithm")

        start_time = time.time()
        if algorithm == 'RC4':
            decrypted_data, decryption_time = handler.decrypt(encrypted_data)
        else:
            decrypted_data, decryption_time = handler.decrypt(encrypted_data, iv)
        
        log_crypto_operation(
            user_id=user_id, file_id=file_record.id, operation_type='decryption',
            algorithm=algorithm, file_size=file_record.file_size,
            execution_time=decryption_time, success=True
        )
        
        response = make_response(decrypted_data)
        response.headers['Content-Type'] = 'application/octet-stream'
        response.headers['Content-Disposition'] = f'attachment; filename="{file_record.original_filename}"'
        response.headers['Content-Length'] = str(len(decrypted_data))
        response.headers['Cache-Control'] = 'no-cache'
        return response
    except Exception as e:
        log_crypto_operation(
            user_id=user_id, file_id=file_record.id, operation_type='decryption',
            algorithm=file_record.encryption_algorithm, file_size=file_record.file_size,
            execution_time=0, success=False, error_message=str(e)
        )
        raise e

def decrypt_file_data_v2(file_record, raw_file_key, user_id):
    """
    Versi V2: Mendekripsi file fisik menggunakan Raw Key yang sudah didapatkan.
    Fungsi ini menerima kunci mentah (bytes) yang sudah didekripsi dari RSA,
    sehingga tidak perlu lagi melakukan derivasi password (PBKDF2).
    """
    try:
        # 1. Baca file terenkripsi dari penyimpanan disk
        # Menggunakan helper 'read_encrypted_file' yang sudah ada di utils/file_handler.py
        encrypted_data = read_encrypted_file(file_record.encrypted_filename)
        
        # 2. Siapkan parameter dekripsi
        # Mengambil IV dari database jika mode cipher memerlukannya (CBC)
        iv = bytes.fromhex(file_record.iv) if file_record.iv else None
        algorithm = file_record.encryption_algorithm

        # 3. Inisialisasi Handler Enkripsi dengan RAW KEY
        # Kita memotong panjang kunci sesuai spesifikasi algoritma jika perlu
        if algorithm == 'AES': 
            handler = AESHandler(raw_file_key) # AES menggunakan full 32 bytes (256 bit)
        elif algorithm == 'DES': 
            handler = DESHandler(raw_file_key[:8]) # DES hanya menggunakan 8 bytes (64 bit)
        elif algorithm == 'RC4': 
            handler = RC4Handler(raw_file_key[:16]) # RC4 menggunakan 16 bytes (128 bit)
        else: 
            raise Exception(f"Unknown encryption algorithm: {algorithm}")

        # 4. Proses Dekripsi & Logging Waktu
        # start_time sebenarnya sudah ditangani di dalam handler.decrypt, 
        # tapi kita ambil return value decryption_time-nya.
        
        if algorithm == 'RC4':
            # RC4 adalah Stream Cipher, tidak butuh IV
            decrypted_data, decryption_time = handler.decrypt(encrypted_data)
        else:
            # AES dan DES (Mode CBC) membutuhkan IV
            decrypted_data, decryption_time = handler.decrypt(encrypted_data, iv)
        
        # 5. Log Operasi Berhasil ke Database
        log_crypto_operation(
            user_id=user_id, 
            file_id=file_record.id, 
            operation_type='decryption',
            algorithm=algorithm, 
            file_size=file_record.file_size,
            execution_time=decryption_time, 
            success=True
        )
        
        # 6. Buat Response Flask untuk Download File
        response = make_response(decrypted_data)
        response.headers['Content-Type'] = 'application/octet-stream'
        # Mengatur nama file agar didownload dengan nama aslinya
        response.headers['Content-Disposition'] = f'attachment; filename="{file_record.original_filename}"'
        response.headers['Content-Length'] = str(len(decrypted_data))
        response.headers['Cache-Control'] = 'no-cache'
        
        return response

    except Exception as e:
        # 7. Log Operasi Gagal (jika ada error)
        # Sangat penting untuk audit trail keamanan
        log_crypto_operation(
            user_id=user_id, 
            file_id=file_record.id, 
            operation_type='decryption',
            algorithm=file_record.encryption_algorithm if file_record else 'Unknown', 
            file_size=file_record.file_size if file_record else 0,
            execution_time=0, 
            success=False, 
            error_message=str(e)
        )
        # Lempar error kembali agar bisa ditangkap oleh blok try-except di handle_download
        raise e

# --- RUTE LAINNYA (Tidak Berubah) ---
@files_bp.route('/download-encrypted/<int:file_id>')
@login_required
def download_encrypted(file_id):
    # (Fungsi ini tidak berubah dari kode Anda)
    can_access, file_record = user_can_access_file(file_id, current_user.id)
    if not can_access:
        flash('You do not have permission to access this file', 'error')
        return redirect(url_for('main.dashboard'))
    
    try:
        filename_to_download = None
        download_as_name = None
        if file_record.file_type == 'excel' and file_record.parsed_filename:
            filename_to_download = file_record.parsed_filename
            download_as_name = f"PARSED_{file_record.original_filename}"
        else:
            filename_to_download = file_record.encrypted_filename
            download_as_name = f"ENCRYPTED_{file_record.original_filename}"
        encrypted_data = read_encrypted_file(filename_to_download)
        return send_file(
            io.BytesIO(encrypted_data),
            as_attachment=True,
            download_name=download_as_name,
            mimetype='application/octet-stream'
        )
    except Exception as e:
        flash(f'Error downloading encrypted file: {str(e)}', 'error')
        return redirect(url_for('main.dashboard'))


@files_bp.route('/delete/<int:file_id>', methods=['POST'])
@login_required
def delete(file_id):
    file_record = db.get_or_404(File, file_id)
    if file_record.owner_id != current_user.id:
        flash('You do not have permission to delete this file', 'error')
        return redirect(url_for('files.my_files'))
    
    try:
        from models.log import CryptoLog
        CryptoLog.query.filter_by(file_id=file_record.id).delete()
        FinancialReport.query.filter_by(file_id=file_record.id).delete()
        
        # --- PERUBAHAN DI SINI ---
        # Hapus 'FileShare.query' karena tabel itu akan dihapus
        # FileShare.query.filter_by(file_id=file_record.id).delete() 
        # (Kita akan menangani penghapusan 'UserAccess' secara berbeda jika diperlukan)
        # --- AKHIR PERUBAHAN ---
        
        from utils.file_handler import delete_file
        delete_file(file_record.encrypted_filename)
        if file_record.parsed_filename:
            delete_file(file_record.parsed_filename)
        
        db.session.delete(file_record)
        db.session.commit()
        flash('File deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting file: {str(e)}', 'error')
    return redirect(url_for('main.dashboard'))


def process_excel_file(file_data, file_id, handler):
    # (Fungsi ini tidak berubah dari Langkah 1)
    try:
        try:
            workbook = openpyxl.load_workbook(BytesIO(file_data))
        except Exception as load_error:
            if "zip" in str(load_error).lower():
                raise Exception("Only .xlsx files are supported for data extraction. Old .xls format is not supported.")
            raise load_error
        
        sheet = workbook.active
        revenue = sheet['B2'].value if sheet['B2'].value else 0
        expenses = sheet['B3'].value if sheet['B3'].value else 0
        profit = sheet['B4'].value if sheet['B4'].value else 0
        
        def encrypt_value(value):
            data = str(value).encode('utf-8')
            ciphertext, iv, _ = handler.encrypt(data)
            return ciphertext.hex(), iv.hex() if iv else None
        
        revenue_enc, _ = encrypt_value(revenue)
        expenses_enc, _ = encrypt_value(expenses)
        profit_enc, _ = encrypt_value(profit)
        
        report = FinancialReport(
            file_id=file_id,
            encrypted_revenue=revenue_enc,
            encrypted_expenses=expenses_enc,
            encrypted_profit=profit_enc,
            encryption_algorithm=handler.key_algorithm_name,
            report_date=datetime.utcnow()
        )
        db.session.add(report)
        db.session.commit()
    except ImportError:
        raise Exception("openpyxl not installed. Install with: pip install openpyxl")
    except Exception as e:
        db.session.rollback()
        raise Exception(f"Excel processing error: {str(e)}")

@files_bp.route('/download-template')
@login_required
def download_template():
    # (Fungsi ini tidak berubah dari kode Anda)
    try:
        root_path = current_app.root_path
        template_filename = 'financial_report_template.xlsx'
        template_path = os.path.join(root_path, template_filename)

        if not os.path.exists(template_path):
            try:
                python_executable = sys.executable
                script_path = os.path.join(root_path, 'create_template.py')
                subprocess.run([python_executable, script_path], check=True, capture_output=True, text=True)
            except subprocess.CalledProcessError as e:
                flash(f'Gagal membuat file template secara internal: {e.stderr}', 'error')
                return redirect(url_for('main.dashboard'))
            except FileNotFoundError:
                flash('Error: create_template.py tidak ditemukan di direktori utama.', 'error')
                return redirect(url_for('main.dashboard'))

        return send_from_directory(
            directory=root_path,
            path=template_filename,
            as_attachment=True
        )
    except Exception as e:
        flash(f'Gagal mengunduh template: {str(e)}', 'error')
        return redirect(url_for('main.dashboard'))