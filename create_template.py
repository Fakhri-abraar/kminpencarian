"""
Create a sample financial report Excel template for testing
"""
import openpyxl

# Create a new workbook
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Financial Report Q4 2024"

# Add headers
sheet['A1'] = "Item"
sheet['B1'] = "Amount (USD)"

# Add financial data
sheet['A2'] = "Revenue"
sheet['B2'] = 1000000

sheet['A3'] = "Expenses"
sheet['B3'] = 750000

sheet['A4'] = "Net Profit"
sheet['B4'] = 250000

# Style the headers
from openpyxl.styles import Font, PatternFill

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")

sheet['A1'].font = header_font
sheet['B1'].font = header_font
sheet['A1'].fill = header_fill
sheet['B1'].fill = header_fill

# Set column widths
sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 20

# Save the file
wb.save('financial_report_template.xlsx')
print("Template created: financial_report_template.xlsx")
print("Sample data:")
print(f"   - Revenue: $1,000,000")
print(f"   - Expenses: $750,000")
print(f"   - Net Profit: $250,000")
print("\nYou can now upload this file through the web interface!")
